"""Parsing of the KRA CRSP workbook into normalised, versionable data.

The source workbook is not clean, so this module deliberately keeps both the
original strings (for display/search) and parsed values (for calculation).
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl


VEHICLE_SHEET = "M.Vehicle CRSP"
MOTORCYCLE_SHEET = "Motor Cycles"
MACHINERY_SHEET = "Tractors & Graders"
TEMPLATE_SHEET = "TEMPLATE"

BLOCK_KEYS = [
    "mv_small",
    "mv_large",
    "mv_high_cc",
    "electric",
    "school_bus",
    "prime_mover",
    "trailer",
    "ambulance",
    "motorcycle",
    "special_purpose",
    "heavy_machinery",
]


def clean(value: Any) -> str:
    """Return a stripped, whitespace-normalised string."""
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\xa0", " ").replace("\u200b", "")
    return re.sub(r"\s+", " ", text).strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def fuel_class(value: Any) -> str:
    text = clean(value).upper()
    if not text:
        return ""
    if "DIESEL" in text:
        return "diesel"
    if "ELECTRIC" in text or re.search(r"\bEV\b", text):
        return "electric"
    if "HYBRID" in text or "PETROL/ELECTRIC" in text:
        return "hybrid"
    if "GASOLINE" in text or "PETROL" in text:
        return "petrol"
    if "CNG" in text or "LNG" in text or "GAS" in text:
        return "gas"
    return ""


BODY_CLASS_MAP = [
    ("AMBULANCE", "ambulance"),
    ("BUS", "bus"),
    ("MINIVAN", "van"),
    ("MINVAN", "van"),
    ("PEOPLE MOVER", "van"),
    ("VAN", "van"),
    ("SUV", "suv"),
    ("CROSSOVER", "suv"),
    ("SEDAN", "sedan"),
    ("SALOON", "sedan"),
    ("SAL", "sedan"),
    ("WAGON", "wagon"),
    ("STATION WAGON", "wagon"),
    ("HATCHBACK", "hatchback"),
    ("HATCBACK", "hatchback"),
    ("COUPE", "coupe"),
    ("CONVERTIBLE", "convertible"),
    ("CONVRTIBLE", "convertible"),
    ("ROADSTER", "convertible"),
    ("SINGLE CAB", "pickup"),
    ("S/CAB", "pickup"),
    ("PICK UP", "pickup"),
    ("PICKUP", "pickup"),
    ("DUAL CAB", "pickup"),
    ("DOUBLE CAB", "pickup"),
    ("CREW CAB", "pickup"),
    ("D/CAB", "pickup"),
    ("TRUCK", "truck"),
    ("TRK", "truck"),
    ("TIPPER", "truck"),
    ("MIXER", "truck"),
    ("TRACTOR", "machinery"),
]


def body_class(value: Any) -> str:
    text = clean(value).upper()
    if not text:
        return ""
    for token, cls in BODY_CLASS_MAP:
        if token in text:
            return cls
    if "MOVER" in text or text == "PM":
        return "prime_mover"
    if re.fullmatch(r"\d+", text):
        return ""
    return ""


def _first_number(text: str) -> float | None:
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def parse_engine(value: Any) -> dict[str, float | None]:
    raw = clean(value)
    if not raw:
        return {"engine_cc": None, "engine_hp": None, "engine_kwh": None, "engine_kw": None}
    upper = raw.upper()
    num = _first_number(raw)
    if "HP" in upper or "PS" in upper or "BHP" in upper:
        return {"engine_cc": None, "engine_hp": num, "engine_kwh": None, "engine_kw": None}
    if "KWH" in upper:
        return {"engine_cc": None, "engine_hp": None, "engine_kwh": num, "engine_kw": None}
    if "KW" in upper:
        return {"engine_cc": None, "engine_hp": None, "engine_kwh": None, "engine_kw": num}
    if num is not None:
        return {"engine_cc": num, "engine_hp": None, "engine_kwh": None, "engine_kw": None}
    return {"engine_cc": None, "engine_hp": None, "engine_kwh": None, "engine_kw": None}


def parse_seating(value: Any) -> int | None:
    num = _first_number(clean(value))
    return int(num) if num is not None else None


def parse_gvw(value: Any) -> tuple[str, float | None]:
    text = clean(value).upper().replace("KG", "").replace("KGS", "").strip()
    number = _to_float(text.replace(" ", ""))
    return clean(value), number


def _model_hint(model: str, body: str) -> str:
    """Body type is missing/misaligned for some rows; use the model text."""
    if body:
        return body
    upper = model.upper()
    for token, cls in (
        ("AMBULANCE", "ambulance"),
        ("TRACTOR", "machinery"),
        ("PRIME MOVER", "prime_mover"),
        ("TRAILER", "trailer"),
        ("TRUCK", "truck"),
        ("BUS", "bus"),
        ("PICKUP", "pickup"),
    ):
        if token in upper:
            return cls
    return ""


def _parse_number_row(row: tuple) -> dict[str, Any]:
    """Normalise a catalogue row that already uses the sheet's canonical columns."""
    make, model = clean(row[0]), clean(row[1])
    body_raw = clean(row[6])
    fuel_raw = clean(row[9])
    engine = parse_engine(row[5])
    gvw_raw, gvw_kg = parse_gvw(row[7])
    return {
        "make": make,
        "model": model,
        "model_number": clean(row[2]),
        "transmission": clean(row[3]),
        "drive": clean(row[4]),
        "engine_raw": clean(row[5]),
        **engine,
        "body_raw": body_raw,
        "body_class": _model_hint(model, body_class(body_raw) or body_class(row[6])),
        "gvw_raw": gvw_raw,
        "gvw_kg": gvw_kg,
        "seating": parse_seating(row[8]),
        "fuel_raw": fuel_raw,
        "fuel_class": fuel_class(fuel_raw),
        "crsp": _to_float(row[10]),
    }


def _header_map(header_row: tuple) -> dict[str, int]:
    aliases = {
        "make": ("MAKE",),
        "model": ("MODEL",),
        "model_number": ("MODELNUMBER", "MODELNO", "MODEL NUMBER"),
        "transmission": ("TRANSMISSION",),
        "drive": ("DRIVECONFIGURATION", "DRIVECONFIG", "DRIVE"),
        "engine": ("ENGINECAPACITY", "ENGINE", "ENGINECAPACITY(CC)"),
        "body": ("BODYTYPE", "BODY"),
        "gvw": ("GVW", "GROSSVEHICLEWEIGHT"),
        "seating": ("SEATING", "SEATS"),
        "fuel": ("FUEL",),
        "crsp": ("CRSP(KES)", "CRSPKES", "CRSP (KES)", "CRSP"),
    }
    result: dict[str, int] = {}
    for index, raw in enumerate(header_row):
        key = re.sub(r"\s+", "", clean(raw)).upper()
        for field_name, candidates in aliases.items():
            if key and key in candidates and field_name not in result:
                result[field_name] = index
    return result


def _parse_catalogue_rows(
    sheet_rows: list[tuple], warnings: list[str], category: str
) -> list[dict[str, Any]]:
    if not sheet_rows:
        return []
    header_map = _header_map(sheet_rows[1])
    rows: list[dict[str, Any]] = []
    for absolute_index, row in enumerate(sheet_rows[2:], start=3):
        if not any(v is not None and clean(v) for v in row):
            continue
        def cell(field: str, default_index: int | None = None):
            index = header_map.get(field, default_index if default_index is not None else -1)
            return row[index] if 0 <= index < len(row) else None

        make, model = clean(cell("make", 0)), clean(cell("model", 1))
        if not make and not model:
            continue
        engine = parse_engine(cell("engine", 5))
        body_raw = clean(cell("body", 6))
        fuel_raw = clean(cell("fuel", 9))
        gvw_raw, gvw_kg = parse_gvw(cell("gvw", 7))
        crsp = _to_float(cell("crsp", 10))
        if crsp is None:
            warnings.append(f"{category} row {absolute_index}: missing CRSP value ({make} {model}).")
        rows.append(
            {
                "category": category,
                "make": make,
                "model": model,
                "model_number": clean(cell("model_number", 2)),
                "transmission": clean(cell("transmission", 3)),
                "drive": clean(cell("drive", 4)),
                "engine_raw": clean(cell("engine", 5)),
                **engine,
                "body_raw": body_raw,
                "body_class": _model_hint(model, body_class(body_raw)),
                "gvw_raw": gvw_raw,
                "gvw_kg": gvw_kg,
                "seating": parse_seating(cell("seating", 8)),
                "fuel_raw": fuel_raw,
                "fuel_class": fuel_class(fuel_raw),
                "crsp": crsp,
                "source_row": absolute_index,
            }
        )
    return rows


def _parse_machinery_rows(sheet_rows: list[tuple], warnings: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_make = ""
    for absolute_index, row in enumerate(sheet_rows[2:], start=3):
        if not any(v is not None and clean(v) for v in row):
            continue
        first, second, third = (clean(row[0]) if len(row) > 0 else ""), (
            clean(row[1]) if len(row) > 1 else ""
        ), (clean(row[2]) if len(row) > 2 else "")
        if first and not second and not third:
            current_make = first
            continue
        if second.upper() == "KSHS":
            continue
        if not first or not third:
            continue
        spec = second if second else None
        crsp = _to_float(third)
        if crsp is None:
            warnings.append(f"Machinery row {absolute_index}: missing CRSP value ({first}).")
        num = _first_number(spec or "") if spec else None
        rows.append(
            {
                "category": "machinery",
                "make": current_make or first,
                "model": first,
                "model_number": "",
                "transmission": "",
                "drive": "",
                "engine_raw": spec or "",
                "engine_cc": None,
                "engine_hp": num,
                "engine_kwh": None,
                "engine_kw": None,
                "body_raw": "",
                "body_class": "machinery",
                "gvw_raw": "",
                "gvw_kg": None,
                "seating": None,
                "fuel_raw": "",
                "fuel_class": "",
                "crsp": crsp,
                "source_row": absolute_index,
            }
        )
    return rows


def _extract_initial_and_divisors(formula: str) -> tuple[float, list[float]]:
    """Parse KRA's customs-value formula into initial retail divisor and back-out divisors.

    Template shape: =((PRICE/1.25)*(100%-DEP)/D1/D2/...)*(100%-EXTRA)
    """
    text = formula.lstrip("=").strip()
    match = re.search(r"/(\d+(?:\.\d+)?)\)?\s*\*?\s*\(?100%", text)
    if not match:
        raise ValueError(f"Unrecognised customs value formula: {formula}")
    initial = float(match.group(1))
    divisors = [float(v) for v in re.findall(r"/(\d+(?:\.\d+)?)", text[match.end() :])]
    if not divisors:
        raise ValueError(f"Unrecognised customs value formula: {formula}")
    return initial, divisors


def _rate_from_value(value: Any, context: str) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value)
    if text.startswith("="):
        percent = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
        if percent:
            return float(percent.group(1)) / 100.0
        fraction = re.search(r"(\d+(?:\.\d+)?)\s*/\s*100", text)
        if fraction:
            return float(fraction.group(1)) / 100.0
    raise ValueError(f"Could not read {context}: {value!r}")


def _map_block_key(title: str, index: int) -> tuple[str, list[str]]:
    upper = title.upper()
    errors: list[str] = []
    if "NOT EXCEEDING 1500" in upper:
        return "mv_small", errors
    if ("8703.24.90" in upper or "8703.33.90" in upper) and "EXCLUDING 8703" not in upper:
        return "mv_high_cc", errors
    if "100% ELECTRIC" in upper or "ELECTRIC POWERED" in upper:
        return "electric", errors
    if "SCHOOL BUSES" in upper:
        return "school_bus", errors
    if "PRIME MOVERS" in upper:
        return "prime_mover", errors
    if "TRAILERS" in upper:
        return "trailer", errors
    if "AMBULANCE" in upper:
        return "ambulance", errors
    if "MOTOR CYCLES" in upper:
        return "motorcycle", errors
    if "SPECIAL PURPOSE" in upper:
        return "special_purpose", errors
    if "HEAVY MACHINERIES" in upper or "HEAVY MACHINERY" in upper:
        return "heavy_machinery", errors
    if "EXCEEDING 1500" in upper:
        return "mv_large", errors
    fallback = BLOCK_KEYS[index] if index < len(BLOCK_KEYS) else None
    errors.append(f"Template block at row is not recognised ({title[:60]}).")
    if fallback:
        return fallback, errors
    raise ValueError(errors[0])


def _parse_template(workbook: Any, warnings: list[str], errors: list[str]) -> dict[str, Any]:
    template_name = find_template_sheet(workbook)
    if template_name is None:
        errors.append("Missing TEMPLATE sheet.")
        return {"blocks": [], "depreciation": {"direct": [], "registered": []}}

    ws = workbook[template_name]
    direct_schedule: list[dict[str, Any]] = []
    registered_schedule: list[dict[str, Any]] = []
    for label_col, rate_col, route in ((1, 2, "direct"), (8, 9, "registered")):
        schedule: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=3, max_row=18):
            label = row[label_col].value
            rate = row[rate_col].value
            if label is None or rate is None:
                continue
            label_text = clean(label).lower()
            if route == "direct":
                nums = re.findall(r"\d+(?:\.\d+)?", label_text)
                if len(nums) >= 2:
                    schedule.append(
                        {
                            "low": float(nums[0]),
                            "high": float(nums[1]),
                            "rate": float(rate),
                            "label": clean(label),
                        }
                    )
            else:
                if label_text.startswith("over"):
                    schedule.append({"age": 999, "rate": float(rate), "label": clean(label)})
                    continue
                nums = re.findall(r"\d+(?:\.\d+)?", label_text)
                if nums:
                    schedule.append({"age": float(nums[0]), "rate": float(rate), "label": clean(label)})
        if route == "direct":
            direct_schedule = schedule
        else:
            registered_schedule = schedule

    blocks: list[dict[str, Any]] = []
    starts: list[int] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[1].value and clean(row[1].value).upper().startswith("TABULATION FOR"):
            starts.append(row[0].row)

    for index, start in enumerate(starts):
        key, block_errors = _map_block_key(str(ws.cell(start, 2).value), index)
        errors.extend(block_errors)
        block: dict[str, Any] = {
            "key": key,
            "title": clean(ws.cell(start, 2).value),
            "duty_rate": None,
            "excise_rate": None,
            "excise_fixed": None,
            "vat_rate": None,
            "rdl_rate": None,
            "idf_rate": None,
            "initial_divisor": None,
            "backout_divisors": [],
        }
        for offset in range(1, 18):
            row_idx = start + offset
            label = clean(ws.cell(row_idx, 2).value).upper()
            value = ws.cell(row_idx, 4).value
            if label.startswith("CUSTOMS VALUE") and value:
                try:
                    block["initial_divisor"], block["backout_divisors"] = _extract_initial_and_divisors(
                        str(value)
                    )
                except ValueError as exc:
                    errors.append(str(exc))
            elif label.startswith("IMPORT DUTY") and value is not None:
                try:
                    block["duty_rate"] = _rate_from_value(value, f"import duty for {key}")
                except ValueError as exc:
                    errors.append(str(exc))
            elif label.startswith("EXCISE DUTY") and value is not None:
                if isinstance(value, (int, float)) and "%" not in clean(ws.cell(row_idx, 2).value):
                    if float(value) == 0:
                        block["excise_rate"] = 0.0
                    else:
                        block["excise_fixed"] = float(value)
                else:
                    try:
                        block["excise_rate"] = _rate_from_value(value, f"excise duty for {key}")
                    except ValueError as exc:
                        errors.append(str(exc))
            elif label.startswith("VAT") and not label.startswith("VAT VALUE") and value is not None:
                try:
                    block["vat_rate"] = _rate_from_value(value, f"VAT for {key}")
                except ValueError as exc:
                    errors.append(str(exc))
            elif label == "RDL" and value is not None:
                try:
                    block["rdl_rate"] = _rate_from_value(value, f"RDL for {key}")
                except ValueError as exc:
                    errors.append(str(exc))
            elif label.startswith("IDF") and value is not None:
                try:
                    block["idf_rate"] = _rate_from_value(value, f"IDF for {key}")
                except ValueError as exc:
                    errors.append(str(exc))

        missing = [name for name, val in (
            ("duty_rate", block["duty_rate"]),
            ("vat_rate", block["vat_rate"]),
            ("initial_divisor", block["initial_divisor"]),
            ("backout_divisors", block["backout_divisors"]),
        ) if val is None or (name == "backout_divisors" and not val)]
        if block["excise_rate"] is None and block["excise_fixed"] is None:
            missing.append("excise")
        if block["rdl_rate"] is None:
            missing.append("rdl")
        if block["idf_rate"] is None:
            missing.append("idf")
        if missing:
            errors.append(f"Template block {key}: missing {', '.join(missing)}.")
        blocks.append(block)

    return {
        "blocks": blocks,
        "depreciation": {"direct": direct_schedule, "registered": registered_schedule},
    }


def find_template_sheet(workbook: Any) -> str | None:
    for name in workbook.sheetnames:
        if "TEMPLATE" in name.upper():
            return name
    return None


def _find_sheet(workbook: Any, prefix: str) -> str | None:
    for name in workbook.sheetnames:
        upper = name.upper()
        if any(token in upper for token in prefix.upper().split(",")):
            return name
    return None


def infer_effective_date(label: str) -> str:
    match = re.search(r"([A-Za-z]+)\s*(\d{4})", label)
    if not match:
        return date.today().isoformat()
    months = {
        "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "MAY": 5,
        "JUNE": 6, "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10,
        "NOVEMBER": 11, "DECEMBER": 12,
    }
    month = months.get(match.group(1).upper())
    year = int(match.group(2))
    if not month:
        return date.today().isoformat()
    return date(year, month, 1).isoformat()


def release_label_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    match = re.search(
        r"(?i)(january|february|march|april|may|june|july|august|september|october|november|december)"
        r"[\w.]*\s*[-.]?\s*(\d{4})",
        stem,
    )
    if match:
        month = match.group(1).capitalize()
        return f"CRSP {month} {match.group(2)}"
    return clean(stem) or "CRSP release"


def parse_workbook(
    source: str | Path | BinaryIO,
    source_filename: str = "",
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    """Parse a KRA CRSP workbook into structured, versionable data."""
    all_warnings: list[str] = warnings if warnings is not None else []
    errors: list[str] = []

    if isinstance(source, (str, Path)):
        values_workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        formula_workbook = openpyxl.load_workbook(source, read_only=False, data_only=False)
    else:
        source.seek(0)
        raw = source.read()
        stream = io.BytesIO(raw)
        values_workbook = openpyxl.load_workbook(stream, read_only=True, data_only=True)
        formula_stream = io.BytesIO(raw)
        formula_workbook = openpyxl.load_workbook(formula_stream, read_only=False, data_only=False)

    label = release_label_from_filename(source_filename or getattr(source, "name", ""))
    effective_date = infer_effective_date(label)

    vehicles: list[dict[str, Any]] = []
    motorcycles: list[dict[str, Any]] = []
    machinery: list[dict[str, Any]] = []

    vehicle_sheet = _find_sheet(values_workbook, "M.Vehicle,CRSP")
    motorcycle_sheet = _find_sheet(values_workbook, "Motor Cycles")
    machinery_sheet = _find_sheet(values_workbook, "Tractors & Graders")
    template_name = find_template_sheet(values_workbook)

    if vehicle_sheet:
        sheet_rows = list(values_workbook[vehicle_sheet].iter_rows(values_only=True))
        vehicles = _parse_catalogue_rows(sheet_rows, all_warnings, "vehicle")
    else:
        errors.append("Could not find the motor vehicle CRSP sheet.")

    if motorcycle_sheet:
        sheet_rows = list(values_workbook[motorcycle_sheet].iter_rows(values_only=True))
        motorcycles = _parse_catalogue_rows(sheet_rows, all_warnings, "motorcycle")
    else:
        errors.append("Could not find the Motor Cycles CRSP sheet.")

    if machinery_sheet:
        sheet_rows = list(values_workbook[machinery_sheet].iter_rows(values_only=True))
        machinery = _parse_machinery_rows(sheet_rows, all_warnings)
    else:
        errors.append("Could not find the Tractors & Graders sheet.")

    template_data = _parse_template(formula_workbook, all_warnings, errors)
    values_workbook.close()
    formula_workbook.close()

    # Ensure the blocks are returned in the canonical order even if KRA
    # reorders the tabulations in a future release.
    block_by_key = {block["key"]: block for block in template_data["blocks"]}
    ordered_blocks = []
    for key in BLOCK_KEYS:
        if key in block_by_key:
            ordered_blocks.append(block_by_key[key])

    parsed = {
        "release_label": label,
        "effective_date": effective_date,
        "vehicles": vehicles,
        "motorcycles": motorcycles,
        "machinery": machinery,
        "blocks": ordered_blocks,
        "depreciation": template_data["depreciation"],
        "counts": {
            "vehicles": len(vehicles),
            "motorcycles": len(motorcycles),
            "machinery": len(machinery),
            "blocks": len(ordered_blocks),
        },
        "warnings": all_warnings,
        "errors": errors,
    }
    return parsed


@dataclass
class CatalogueRow:
    """A single row fetched from the database, used by the search/calculation layer."""

    id: int = 0
    release_id: int = 0
    category: str = ""
    make: str = ""
    model: str = ""
    model_number: str = ""
    transmission: str = ""
    drive: str = ""
    engine_raw: str = ""
    engine_cc: float | None = None
    engine_hp: float | None = None
    engine_kwh: float | None = None
    engine_kw: float | None = None
    body_raw: str = ""
    body_class: str = ""
    fuel_raw: str = ""
    fuel_class: str = ""
    seating: int | None = None
    crsp: float | None = None
    source_row: int = 0

    def display_name(self) -> str:
        parts = [self.make, self.model]
        if self.model_number:
            parts.append(f"({self.model_number})")
        return " ".join(parts)

    def spec_label(self) -> str:
        bits = []
        if self.engine_raw:
            bits.append(self.engine_raw)
        if self.fuel_raw:
            bits.append(self.fuel_raw)
        if self.body_raw:
            bits.append(self.body_raw)
        if self.seating:
            bits.append(f"{self.seating} seats")
        return " · ".join(bits)

    def as_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "category": self.category,
            "make": self.make,
            "model": self.model,
            "model_number": self.model_number,
            "display": self.display_name(),
            "spec": self.spec_label(),
            "engine_cc": self.engine_cc,
            "fuel_class": self.fuel_class,
            "body_class": self.body_class,
            "crsp": round(self.crsp, 2) if self.crsp is not None else None,
        }


def row_to_catalogue(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "category": row.get("category", ""),
        "make": row.get("make", ""),
        "model": row.get("model", ""),
        "model_number": row.get("model_number", ""),
        "transmission": row.get("transmission", ""),
        "drive": row.get("drive", ""),
        "engine_raw": row.get("engine_raw", ""),
        "engine_cc": row.get("engine_cc"),
        "engine_hp": row.get("engine_hp"),
        "engine_kwh": row.get("engine_kwh"),
        "engine_kw": row.get("engine_kw"),
        "body_raw": row.get("body_raw", ""),
        "body_class": row.get("body_class", ""),
        "gvw_raw": row.get("gvw_raw", ""),
        "gvw_kg": row.get("gvw_kg"),
        "seating": row.get("seating"),
        "fuel_raw": row.get("fuel_raw", ""),
        "fuel_class": row.get("fuel_class", ""),
        "crsp": row.get("crsp"),
        "source_row": row.get("source_row", 0),
    }
